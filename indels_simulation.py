import numpy as np
import math
from scipy import stats
from scipy.optimize import fsolve
from numpy import *
import warnings
import sys
import os


class Node:
  def __init__(self):
    self.name = ""
    self.children = []
    self.lengthS = ""
    self.length = 0
    self.parent = None
    self.genom = []
    self.distances = []
    self.genomLength = []
class Tree:
    def __init__(self):
        self.leaves = []
        self.root = None
        self.total = []
    def getLeaves (self):
        for x in range(len(self.leaves)):
            print(self.leaves[x].name)
    def getRandomLeaf (self):
        r1 = random.randint(0,len(self.leaves) -1)
        return r1


#------------------------------------------------------

next_number=5001

#function to read the rf value, and to normalize it.
#get = name of file
#return = rf, normalized rf
def findRF(name):
    temp = open("treedist" + str(name) + ".txt",'r')
    f = temp.read()
    #get rf 
    RF = f.lstrip('1')
    first = open ("intree", 'r')
    first = first.read()
    second = open ("second" + str(name) + ".txt", 'r')
    second = second.read()
    count1 = first.count("(") - 1 - binaryTree(first)
    #count1 = first.count("(") 
    count2 = second.count("(") - 1 - binaryTree(second)  
    #common_edges = (count1 + count2 - int(RF))/2
    norm = int(RF)/ (count1 + count2)
    #result = tree distance
    return int(RF), norm

#main function to check if the tree is binary
#get=string (newick)
#return = true or false
def binaryTree (strTree):
 tree=Tree()
 root, count = createTree (strTree, tree, 0, None)
 tree.root = root
 return checkBinary(tree.root)

#recursive function to check if tree is binary.
#get the root of the tree.
#return true or false
def checkBinary(node) :
  if(len(node.children) == 0):
     return True
  if(len(node.children) == 1 or (len(node.children) >2)):
      return False
  for x in range (len(node.children)):
      if not checkBinary(node.children[x]):
        return False
  return True 

#create new tree structure
#get : nwstr = newick format string, tree = instance of Tree object, count = pointer to char in the string 'nwstr', prnt = Node's parent to be, in the first call its None (because root doesnt have parent)
#returns : at each iteration returns the last node we finished creating, last iteration : the root of the tree;  count = pointer to char in the string 'nwstr'                    
def createTree (nwstr,tree, count, prnt):
    while (len(nwstr) > count and (nwstr[count] == '\n' or nwstr[count] == ' ')):
         count+=1
    #if we at the end of the string, return      
    if (len(nwstr) == count):
      return None, count
    #if we see "(" , its a new node
    if (nwstr[count] == "("):
       root = Node()
       root.parent= prnt
       while (len(nwstr) > count and (nwstr[count] == '\n' or nwstr[count] == ' ')):
         count+=1
       count+=1
       #call function again, the func will return the child of that node
       child, count = createTree (nwstr,tree, count, root)
       root.children.append(child)
       #while wee see "," , there are more children  
       while (nwstr[count] == ','):
         while (len(nwstr) > count and (nwstr[count] == '\n' or nwstr[count] == ' ')):
           count+=1
         count+=1
        #call function again, the func will return the child of that node
         child, count = createTree (nwstr,tree, count, root)
         root.children.append(child)
       count+=1
       while (len(nwstr) > count and (nwstr[count] == '\n' or nwstr[count] == ' ')):
         count+=1
       #if we at the end of the string, return      
       if (len(nwstr) <= count):
        return root,count
       else :
        if (nwstr[count]  == ":"):
          count +=1
        #we want to read the name, we search till we see  ":", knowing after that comes the length  
        else :
          while (count < len (nwstr) and nwstr[count]  != ":"):
            while (len(nwstr) > count and (nwstr[count] == '\n' or nwstr[count] == ' ')):
               count+=1
            if (len(nwstr) <= count ):
              break
            root.name += nwstr[count]
            count+=1
          count+=1
        while (len(nwstr) > count and (nwstr[count] == '\n' or nwstr[count] == ' ')):
           count+=1
        #we read the length, stoped when we see "," or ")"  
        while (len(nwstr) > count and (nwstr[count] != "," and nwstr[count] != ")")):
          while (len(nwstr) > count and (nwstr[count] == '\n' or nwstr[count] == ' ')):
               count+=1
          root.lengthS +=nwstr[count]
          count+=1
       return root,count
    #if we dont see "(", the new node is a leaf
    else:
        leaf = Node()
        leaf.parent = prnt
        while (len(nwstr) > count and (nwstr[count] == '\n' or nwstr[count] == ' ')):
           count+=1
        if (nwstr[count]  == ":"):
          count +=1
        else :
        #we want to read the name, we search till we see  ":", knowing after that comes the length  
          while (nwstr[count]  != ":"):
            while (nwstr[count] == '\n' or nwstr[count] == ' '):
                count+=1
            leaf.name += nwstr[count]
            count+=1
          count+=1
       #we read the length, stoped when we see "," or ")"    
        while (len(nwstr) > count and (nwstr[count] != "," and nwstr[count] != ")")):
          while (nwstr[count] == '\n' or nwstr[count] == ' '):
                count+=1
          leaf.lengthS +=nwstr[count]
          count+=1
        #add the new leaf to tree.leaves  
        tree.leaves.append(leaf)
        return leaf, count


#function to create new newick tree from given distances matrix (read from file).
#the function also calls for treedist with the new newick and the base newick("first"+name).
#get = name of file (distance matrix).
#created two relevant files : "second"+name(new newick tree), "treedist"+name (contains rf).
def createIntree2(name):
   if os.path.exists("intree2"):
      os.remove("intree2")
   if os.path.exists("outfile"):
      os.remove("outfile")
   if os.path.exists("outtree"):
      os.remove("outtree")   
   os.system("echo y| phylip neighbor")
   #rename files 
   old_name = r"outfile"
   new_name = r"new_tree" + str(name) + ".txt"
   if os.path.exists(new_name):
    os.remove(new_name)
   os.rename(old_name, new_name)
   old_name = r"outtree"
   new_name =  r"intree2"
   os.rename(old_name, new_name)
   #call treedist
   stat = os.system("phylip treedist < phylip-param.txt > dist.out.txt"); 
   #rename files 
   old_name = r"outfile"
   new_name = r"treedist" + str(name) + ".txt"
   if os.path.exists(new_name):
    os.remove(new_name)
   os.rename(old_name, new_name)
   old_name = r"intree2"
   new_name = r"second" + str(name) + ".txt"
   if os.path.exists(new_name):
    os.remove(new_name)
   os.rename(old_name, new_name)
   
#func to create string of newick format from tree 
def createNewickFormat (root,newick):
    if (len(root.children) == 0):
        if (root.length != 0):        
            newick += str(root.name) + ":" + str("{:.3f}".format(root.length))
            return newick
    else:
        newick += "("
        newick = createNewickFormat(root.children[0], newick)
        newick += ","
        newick = createNewickFormat(root.children[1], newick)
        newick += ")"
        if (root.length != 0):
           newick+= ":" + str("{:.3f}".format(root.length))
    return newick

#function to create genome from given one
#get : two (same) genes of the parent, number of jumps    
#return : new genom    
def change_vector_indels (changed_gen, change, ratio):              
    x = 0
    #num = max(changed_gen) + 1
    while x < change :
        randNumber = random.random()
        #case of jump :
        if (randNumber < ratio):
          randNumber = random.random()
          if (randNumber <= 0.5):
            y = random.randint(len(changed_gen))
            z = random.randint(len(changed_gen))
            element = changed_gen.pop(y)
            changed_gen.insert (z,element)
        else :
           randNumber = random.random()
           #case of insert
           if (randNumber <0.5):
            z = random.randint(len(changed_gen))
            global next_number
            changed_gen.insert (z,next_number)
            next_number+=1
           # num +=1
           else:
              y = random.randint(len(changed_gen))
              changed_gen.pop(y)
        x+=1
    return changed_gen
    
#create tree (binary) 
#get = tree(empty at first), n = number of leaves, scaleexp = input for length of branch, size of Genom, indels percentage.
def create_recr_tree_rand (tree, n, scaleexp ,sizeofGenom, indels):                               
    if (len(tree.leaves) == n):
        return
    if (len(tree.leaves) == 0):
        root = Node()
        size = sizeofGenom
        for x in range(size):
            root.genom.append(x)
        tree.root = root
        tree.leaves = [root]
        tree.total = [root]
        create_recr_tree_rand (tree, n, scaleexp ,sizeofGenom, indels)
    else:
        if (len(tree.leaves) == 1):
                numLeaf = 0
        else:        
             numLeaf = tree.getRandomLeaf() 
        chosenLeaf = tree.leaves[numLeaf]
        left = Node()
        left.length = random.exponential(scale=scaleexp)
        left.parent = chosenLeaf
        change = random.poisson(lam=(sizeofGenom*2+1)*left.length, size=1)
        left.genom= change_vector_indels ( left.parent.genom.copy(), change, indels)
        right = Node()
        right.length = random.exponential(scale=scaleexp)
        right.parent = chosenLeaf
        change = random.poisson(lam=(sizeofGenom*2+1)*right.length, size=1)
        right.genom = change_vector_indels (right.parent.genom.copy(), change, indels)
        chosenLeaf.children = [left,right]
        tree.leaves.pop(numLeaf)
        tree.leaves.append(left)
        tree.leaves.append(right)
        tree.total.append(left)
        tree.total.append(right)
        create_recr_tree_rand (tree, n, scaleexp ,sizeofGenom, indels)
        
#function to create new tree, add names for leaves, create newick file from this tree (with name intree).
#get = tree(empty at first), n = number of leaves, scaleexp = input for length of branch, size of Genom, indels percentage.
def create_intree(tree, n, scaleexp ,sizeofGenom, indels):
   create_recr_tree_rand (tree, n, scaleexp ,sizeofGenom, indels) 
   for x in range (n):
      tree.leaves[x].name = "species_" + str(x) 
   print ("created tree") 
   #export the tree to newick formt to file called "intree"  
   newstr = createNewickFormat(tree.root, "") + ";" 
   f = open("intree", "w")
   f.write(newstr)
   f.close()
   

#k=6 -> si and 1-si
#k=1 -> cga
#k=2 -> LCS
#k=3 -> g_ab
#k=4 -> jaccard
#function to create distance matrix for fiven method.
#get array of genomes and number k (indicates what method to choose).
def create_Matrix (arr,k):
  siLength = [[0]*len(arr) for i in range(len(arr))]  
  siLength_minus = [[0]*len(arr) for i in range(len(arr))]
  if (k==6):
    print("si and 1-si:")
    for x in range (len(arr)):
       for y in range (x, len(arr)):
          if (x == y):
            siLength[x][x] = 0
            siLength_minus[x][x]=0
          else: 
           result = SyntenyIndex_indels(arr[x].genom, arr[y].genom)
           if (result == 1):
                 dist = []
                 dist.append(0)
           else: 
             dist = [] 
             dist = fsolve(calcExactDistFromSi_6, 1.0, args = result)  #1.0 is the initial value for tvalue for t   
           siLength[x][y] = round(dist[0],6)
           siLength[y][x] = round(dist[0],6)
           temp=round(1-result,6) 
           siLength_minus[x][y] = temp
           siLength_minus[y][x] = temp
  elif (k==1):
      print("cga:")
      for x in range (len(arr)):
       for y in range (x, len(arr)):
          if (x == y):
            siLength[x][x] = 0
            siLength_minus[x][x]=0
          else: 
            result=CGA(arr[x].genom, arr[y].genom)
            siLength[x][y] = round(result,6) 
            siLength[y][x] = round(result,6)
  elif(k==2):
      print("lcs:")
      for x in range (len(arr)):
       for y in range (x, len(arr)):
          if (x == y):
            siLength[x][x] = 0
            siLength_minus[x][x]=0
          else: 
            result=LCS(arr[x].genom, arr[y].genom)
            siLength[x][y] = round(result,6) 
            siLength[y][x] = round(result,6)
  elif(k==3):
      print("gab:")
      for x in range (len(arr)):
       for y in range (x, len(arr)):
          if (x == y):
            siLength[x][x] = 0
            siLength_minus[x][x]=0
          else: 
            result=calculate_g_ab(arr[x].genom, arr[y].genom)
            siLength[x][y] = round(result,6) 
            siLength[y][x] = round(result,6)      
  elif(k==4):
      print("jaccard:")
      for x in range (len(arr)):
       for y in range (x, len(arr)):
          if (x == y):
            siLength[x][x] = 0
            siLength_minus[x][x]=0
          else: 
            result=jaccard(arr[x].genom, arr[y].genom)
            siLength[x][y] = round(result,6) 
            siLength[y][x] = round(result,6)             
  newStr = str(len(arr)) + "\n"   
  newStr_minus = str(len(arr)) + "\n"         
  for x in range (len(arr)):
       newStr += str(arr[x].name)
       newStr_minus+= str(arr[x].name)
       for y in range (len(arr)):
            newStr+= " " + str(siLength[x][y])  
            newStr_minus+= " " + str(siLength_minus[x][y])  
       newStr+="\n"  
       newStr_minus+="\n" 
  return newStr, newStr_minus

#calculate si
def SyntenyIndex_indels(Genom1, Genom2, k=6):
    total = (len(Genom1)+len(Genom2))/2
    t = 0
    for item in range(k, len(Genom1) - k):
        if (Genom1[item] in Genom2 and Genom1[item]!=-1):
          set1 = set(Genom1[item - k:item])
          set2 = set(Genom1[item + 1:item + k + 1])
          set3 = set1.union(set2)
          item1 = Genom2.index(Genom1[item])
          set12 = set(Genom2[item1 - k:item1])
          set22 = set(Genom2[item1 + 1:item1 + k + 1])
          set32 = set12.union(set22)
          set3_filtered = {x for x in set3 if x != -1}
          set32_filtered = {x for x in set32 if x != -1}
	# Calculate the intersection of the filtered sets
          intersection = set3_filtered.intersection(set32_filtered)
          len_intersection = len(intersection)
          b = (1 / (2 * k))
          JSI = b * len_intersection
          t += JSI
    if t==0:
      t=1      
    #total = (len(Genom1)+len(Genom2))/2
    syntenyindex = (1 / total) * t
    return syntenyindex

#for k = 6 in SI, calculate the value of the si.
def calcExactDistFromSi_6(t, synIn):
  if (synIn == 0.0):
               print ("fgfgg")
               return(1000.0)
  #if (synIn == 1):
  #        return 0 
  else:
    return(math.exp(2*t)*synIn - (6*t**10 + 30*t**9 + 115*t**8 + 230*t**7 + 362*t**6 + 362*t**5 + 280*t**4 + 140*t**3 + 50*t**2 + 10*t + 1)/(t + 1)**11)
    

def LCS (X,Y):
   arr1=X.copy()
   arr2=Y.copy()
   x=0
   while x<len(arr1):
     if not arr1[x] in arr2 or arr1[x]==-1:
        arr1.pop(x)
     else:
        x+=1  
   x=0
   while x<len(arr2):
     if not arr2[x] in arr1 or arr2[x]==-1:
        arr2.pop(x)
     else:
        x+=1
   if(len(arr1)==0):
     res= 1/((len(X)+len(Y))/2)
     return -math.log(res)

   nums=[None]*len(arr2)
   for x in range (len(nums)):
      indx=arr1.index(arr2[x])
      nums[x]=indx
     
   n = len(nums)
   ans = []
 
    # Initialize the answer list with the
    # first element of nums
   ans.append(nums[0])
 
   for i in range(1, n):
        if nums[i] > ans[-1]:
            # If the current number is greater
            # than the last element of the answer
            # list, it means we have found a
            # longer increasing subsequence.
            # Hence, we append the current number
            # to the answer list.
            ans.append(nums[i])
        else:
            # If the current number is not
            # greater than the last element of
            # the answer list, we perform
            # a binary search to find the smallest
            # element in the answer list that
            # is greater than or equal to the
            # current number.
            low = 0
            high = len(ans) - 1
            while low < high:
                mid = low + (high - low) // 2
                if ans[mid] < nums[i]:
                    low = mid + 1
                else:
                    high = mid
            # We update the element at the
            # found position with the current number.
            # By doing this, we are maintaining
            # a sorted order in the answer list.
            ans[low] = nums[i]
 
    # The length of the answer list
    # represents the length of the
    # longest increasing subsequence.
   lcs=len(ans)
   if(lcs==0):
        lcs=1
   res= lcs/((len(X)+len(Y))/2)
   return -math.log(res)



#calculate xi
def CGA(Genom1, Genom2):
    sharedPairs=0
    for x in range (len(Genom1)-1):
      if (Genom1[x] in Genom2 and Genom1[x]!=-1 and Genom1[x+1]!=-1):
         indx = Genom2.index(Genom1[x])
         if(indx+1!=len(Genom2)):
            if(Genom2[indx+1]==Genom1[x+1]):
               sharedPairs+=1
    L0 = (len(Genom1) + len(Genom2))/ 2
    if (sharedPairs==0):
      sharedPairs=1          
    P00=sharedPairs/(L0-1)
    '''
    try:
      xi=0.5*((1-1/P00**0.5)*log(P00))**0.5
    except:
      xi=LCS(Genom1, Genom2)
      return xi 
    '''
    xi=0.5*((1-1/P00**0.5)*log(P00))**0.5   
    itr=100
    for itern in range(itr):
       xi=-(log(1+xi)+log(P00))/2   
    return xi

#create input file for dcj
def create_dcj_mat (tree):
   mat=''
   for x in range (len(tree.leaves)):
     mat+=">"+str(tree.leaves[x].name) + "\n"
     mat+=(" ".join(map(str, tree.leaves[x].genom)))+ "\n"
   f = open("dcj.txt", "w")
   f.write(mat)
   f.close()
   if os.path.exists("dcj_matrix.txt"):
      os.remove("dcj_matrix.txt")
   os.system("java -jar UniMoG-java11.jar -m=6 -d dcj.txt >>dcj_matrix.txt")

#create input file for hp
def create_hp_mat (tree):
   mat=''
   for x in range (len(tree.leaves)):
     mat+=">"+str(tree.leaves[x].name) + "\n"
     mat+=(" ".join(map(str, tree.leaves[x].genom)))+ "\n"
   f = open("dcj.txt", "w")
   f.write(mat)
   f.close()  
   if os.path.exists("dcj_matrix.txt"):
      os.remove("dcj_matrix.txt")   
   os.system("java -jar UniMoG-java11.jar -m=3 -d dcj.txt >>dcj_matrix.txt")
   
   
#create input file for neighbor from dcj
def read_dcj_mat():
  import re
  with open('dcj_matrix.txt', 'r') as file:
    file_content = file.read()

# Define a regular expression pattern
  pattern = re.compile(r"DCJ-indel distance comparisons as PHYLIP matrix:(.*?)Steps of each genome comparison:", re.DOTALL)

# Find the matched content
  match = re.search(pattern, file_content)
  names=[]
# Extract the content if the pattern is found
  if match:
    extracted_content = match.group(1).strip()
    # Split the content into lines
    lines = extracted_content.split('\n')[1:]
    # Initialize an empty matrix
    matrix = []
    indx=0
    # Iterate through each line and parse the values
    for line in lines:
        values = line.split()
        if not "species_"+str(indx) == values[0]:
          string1 ="species_"+str(indx)
          num=values[0].replace(string1, "") 
          row = [float(num)] + [int(value) if value.isdigit() or (value[0] == '-' and value[1:].isdigit()) else value for value in values[1:]]
        else:
          row = [int(value) if value.isdigit() or (value[0] == '-' and value[1:].isdigit()) else value for value in values[1:]]  
        names.append("species_"+str(indx))
        indx+=1
        # Convert each value to an integer (skip species names)
        matrix.append(row)
    siLength = [[0]*len(lines) for i in range(len(lines))]      
    for i in range(len(lines)):
       for j in range(len(lines)):
          if i < len(matrix) and j < len(matrix[i]):
            siLength[i][j] = round(matrix[i][j],6)
            siLength[j][i] = round(matrix[i][j],6)
    newStr = str(len(lines)) + "\n"   
    for x in range (len(lines)):
       newStr += str(names[x])
       for y in range (len(lines)):
            newStr+= " " + str(siLength[x][y])  
       newStr+="\n"   
    return newStr
#create input file for neighbor from hp
def read_hp_mat():
  import re
  with open('dcj_matrix.txt', 'r') as file:
    file_content = file.read()

# Define a regular expression pattern
  pattern = re.compile(r"HP distance comparisons as PHYLIP matrix:(.*?)Steps of each genome comparison:", re.DOTALL)

# Find the matched content
  match = re.search(pattern, file_content)
  names=[]
# Extract the content if the pattern is found
  if match:
    extracted_content = match.group(1).strip()
    # Split the content into lines
    lines = extracted_content.split('\n')[1:]
    # Initialize an empty matrix
    matrix = []
    indx=0
    # Iterate through each line and parse the values
    for line in lines:
        values = line.split()
        if not "species_"+str(indx) == values[0]:
          string1 ="species_"+str(indx)
          num=values[0].replace(string1, "") 
          row = [float(num)] + [int(value) if value.isdigit() or (value[0] == '-' and value[1:].isdigit()) else value for value in values[1:]]
        else:
          row = [int(value) if value.isdigit() or (value[0] == '-' and value[1:].isdigit()) else value for value in values[1:]]  
        names.append("species_"+str(indx))
        indx+=1
        # Convert each value to an integer (skip species names)
        matrix.append(row)
    siLength = [[0]*len(lines) for i in range(len(lines))]      
    for i in range(len(lines)):
       for j in range(len(lines)):
          if i < len(matrix) and j < len(matrix[i]):
            siLength[i][j] = round(matrix[i][j],6)
            siLength[j][i] = round(matrix[i][j],6)
    newStr = str(len(lines)) + "\n"   
    for x in range (len(lines)):
       newStr += str(names[x])
       for y in range (len(lines)):
            newStr+= " " + str(siLength[x][y])  
       newStr+="\n"   
    return newStr
    
def calculate_g_ab (a,b):
  c_ab = len (set(a).intersection(b))
  m_a = len(np.unique(a))
  m_b = len(np.unique(b))
  if(c_ab==0):
    c_ab=1
  val = c_ab/ min (m_a, m_b)
  return -math.log (val)
 
def jaccard (a,b):
  c_ab = len (set(a).intersection(b))
  if(c_ab==0):
    c_ab=1
  val = c_ab/ (len(a)+len(b) - c_ab)
  return -math.log (val)


#k=0 -> hp
#k=1 -> cga
#k=2 -> LCS
#k=3 -> g_ab
#k=4 -> jaccard
#k=5 -> dcj
#k=6 -> si
#k=7 -> 1-si
#main function to run simulations.
#get the methods it want to run, the name of the files you want, indels = 1.1 for only jump. 0 for only indels. in middle (0-1) : combination, closer to 1 - more jumps.
#l = start mean length of branches, add = by how much the branch length grow by each iteration, n = number of leaves in tree. Gsize=mean size of genome (or exact for only jumps).
#run for 50 repeats for each simulation. run for 10 different lengths.
#if want to change  : change th itr number (for number of repeats), limit number (for number of lengths to check).
#return the array of lengths cheked, matrix off rf results (length X methods), matrix of normalized rf results).
def createTreeSimulationsMultiple (methods, name,indels, l=0.001, add=0.001,n=20, Gsize=5000):
  itr=50
  limit=10
  results=[]
  resultsNorm = []
  arrLen = []
  for i in range (limit):
     if(6 in methods):
       iterResult =  [[] for _ in range(len(methods)+1)]
       iterResultsNorm = [[] for _ in range(len(methods)+1)]
     else:  
      iterResult = [[] for _ in range(len(methods))]
      iterResultsNorm = [[] for _ in range(len(methods))]
     for j in range (itr):
         for z in range (len(methods)):
           if(z==0):
            if(methods[z]==0): 
             success = True
             while (success):
              try:
               tree = Tree()
               create_intree(tree, n, l ,Gsize, indels)
               #hp
               print("hp")
               create_hp_mat(tree)
               mat1=read_hp_mat()
               f = open("infile", "w")
               f.write(mat1)
               f.close()
               createIntree2(name)
               RF,RF_norm = findRF(name)
               iterResult[z].append(RF)
               iterResultsNorm[z].append(RF_norm)
               success=False
              except:
               print("failed")
            else:
              tree = Tree()
              create_intree(tree, n, l ,Gsize, indels)
              if(methods[z]==5):
                print("dcj")
                create_dcj_mat(tree)
                mat1=read_dcj_mat()
              else:
                mat1, mat2 = create_Matrix (tree.leaves,methods[z])
              f = open("infile", "w")
              f.write(mat1)
              f.close()
              createIntree2(name)
              RF,RF_norm = findRF(name)
              iterResult[z].append(RF)
              iterResultsNorm[z].append(RF_norm)
              if(methods[z]==6):
                f = open("infile", "w")
                f.write(mat2)
                f.close()
                createIntree2(name)
                RF,RF_norm = findRF(name)
                iterResult[z+1].append(RF)
                iterResultsNorm[z+1].append(RF_norm)
           else:
              if(methods[z]==5):
                print("dcj")
                create_dcj_mat(tree)
                mat1=read_dcj_mat()
              else:
                mat1, mat2 = create_Matrix (tree.leaves,methods[z])
              f = open("infile", "w")
              f.write(mat1)
              f.close()
              createIntree2(name)
              RF,RF_norm = findRF(name)
              iterResult[z].append(RF)
              iterResultsNorm[z].append(RF_norm)
              if(methods[z]==6):
                f = open("infile", "w")
                f.write(mat2)
                f.close()
                createIntree2(name)
                RF,RF_norm = findRF(name)
                iterResult[z+1].append(RF)
                iterResultsNorm[z+1].append(RF_norm)
     arrLen.append(l)
     l+=add
     results.append(iterResult)
     resultsNorm.append(iterResultsNorm)
  return arrLen,results,resultsNorm
  

#1-cga, 2-lcs, 6-si , 61 - 1-si, 5- dcj, 3 - gab, 4 - jaccard, 7 - hp
#calles for the simulation function.
#creates text file of results, and calculate the mean rf and std from the returned matrix for each method in each length.
def calcMeanAndResidualsForMultiple(methods,name, indels, l, add,n, Gsize):
  arr_l,results,resultsNorm=createTreeSimulationsMultiple(methods,name, indels, l, add,n, Gsize)
  if(6 in methods):
    methods.append(7)
  arr_rf = [[] for _ in range(len(results[0]))]
  arr_std =  [[] for _ in range(len(results[0]))]
  arr_rf_norm = [[] for _ in range(len(results[0]))]
  arr_std_norm =  [[] for _ in range(len(results[0]))]
  method_names = ['hp','cga', 'lcs', 'gab', 'jaccard', 'dcj', 'si', '1-si']
  f = open("restults_"+str(name), "a")
  newstr = "for "+ str(indels)+ "\n"
  for i in range (len(arr_l)):
    newstr += "for len : " + str(arr_l[i])+"\n"
    for j in range (len(results[i])):
      newstr+=str(method_names[methods[j]]) + ": " + str(results[i][j]) + "\n" + str(method_names[methods[j]]) + " norm: " + str(resultsNorm[i][j]) + "\n"
      arr_rf[j].append(mean(results[i][j]))
      arr_std[j].append(np.std(results[i][j], ddof=1))
      arr_rf_norm[j].append(mean(resultsNorm[i][j]))
      arr_std_norm[j].append(np.std(resultsNorm[i][j], ddof=1))
    f.write(newstr)
  f.close()
  return arr_l,arr_rf,arr_std,arr_rf_norm,arr_std_norm
  
  
#creates excel file with the results from the simualtion.
#get the methods it want to run, the name of the files you want, indels = 1.1 for only jump. 0 for only indels. in middle (0-1) : combination, closer to 1 - more jumps.
#l = start mean length of branches, add = by how much the branch length grow by each iteration, n = number of leaves in tree. Gsize=mean size of genome (or exact for only jumps).

def createExcel_for_tree_multiple(methods,name, indels, l = 0.001, add = 0.001,n=20, Gsize=5000):
 import openpyxl
 from openpyxl import Workbook
 global next_number
 next_number=5001
 arr_l,arr_rf,arr_std,arr_rf_norm,arr_std_norm = calcMeanAndResidualsForMultiple(methods,name, indels, l, add,n, Gsize)
 if(6 in methods):
    methods.append(7)
 method_names = ['hp','cga', 'lcs', 'gab', 'jaccard', 'dcj', 'si', '1-si']
 excel_file_path=(str(name)+".xlsx")
 try:
    # Try to load the existing Excel file
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    sheet.cell(row=1, column=sheet.max_column+1, value="for "+ str(indels))
 except FileNotFoundError:
    # If the file doesn't exist, create a new workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Add the new header as the first column
    sheet.insert_cols(1)
    sheet.cell(row=1, column=1, value="for "+ str(indels))
 #run on all methods :
 indx=1
 col=sheet.max_column+1
 sheet.cell(row=1, column=col, value="length")
 for x in range (len(arr_l)):
    indx+=1
    sheet.cell(row=indx, column=col, value=arr_l[x])
 for i in range (len(arr_rf)):  
  indx=1
  col=sheet.max_column+1
  sheet.cell(row=1, column=col, value=method_names[methods[i]])
  for x in range (len(arr_rf[i])):
    indx+=1
    sheet.cell(row=indx, column=col, value=arr_rf[i][x])
 for i in range (len(arr_std)):  
  indx=1
  col=sheet.max_column+1
  sheet.cell(row=1, column=col, value=method_names[methods[i]]+"_std")
  for x in range (len(arr_std[i])):
    indx+=1
    sheet.cell(row=indx, column=col, value=arr_std[i][x])
 for i in range (len(arr_rf_norm)):  
  indx=1
  col=sheet.max_column+1
  sheet.cell(row=1, column=col, value=method_names[methods[i]]+"_norm")
  for x in range (len(arr_rf_norm[i])):
    indx+=1
    sheet.cell(row=indx, column=col, value=arr_rf_norm[i][x])
 for i in range (len(arr_std_norm)):  
  indx=1
  col=sheet.max_column+1
  sheet.cell(row=1, column=col, value=method_names[methods[i]]+"_norm_std")
  for x in range (len(arr_std_norm[i])):
    indx+=1
    sheet.cell(row=indx, column=col, value=arr_std_norm[i][x])
 workbook.save(excel_file_path)
 workbook.close()


#k=0 -> hp
#k=1 -> cga
#k=2 -> LCS
#k=3 -> g_ab
#k=4 -> jaccard
#k=5 -> dcj
#k=6 -> si + 1-si
methods=[1,3,5,6]
createExcel_for_tree_multiple(methods, "methods",0,0.1,0.1) 
